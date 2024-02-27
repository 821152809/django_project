from django.contrib import admin
from django.http import HttpResponse
from vocabulary_management import models
from vocabulary_management.logic.language_trans import LanguageTrans
from import_export.admin import ImportExportModelAdmin
import jieba, openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from django.core.files.storage import FileSystemStorage
import difflib

admin.site.site_header = "飞鹤-数据中心组后台管理系统"
admin.site.index_title = "首页"


@admin.register(models.VocabularyManagement)
class DataProcess(ImportExportModelAdmin):
    # 列表页显示那些字段（列）
    list_display = [
        "vocabulary_en",
        "vocabulary_cn",
        "vocabulary_type",
        "vocabulary_desc",
    ]
    search_fields = [
        "vocabulary_en",
        "vocabulary_cn",
    ]


@admin.register(models.VocabularyTranslate)
class DataProcess(admin.ModelAdmin):
    # 列表页显示那些字段（列）
    list_display = ["vocabulary", "trans_mode", "trans_result", "trans_result_lower"]
    list_filter = ("vocabulary",)
    search_fields = ["vocabulary"]

    def start_trans(modeladmin, request, queryset):
        for obj in queryset:
            if obj.trans_mode == 0:
                trans_mode_in = "E2C"
            else:
                trans_mode_in = "C2E"
            translator = LanguageTrans(trans_mode_in)
            trans_result = translator.trans(obj.vocabulary)
            obj.trans_result = trans_result
            obj.trans_result_lower = trans_result.lower()
            obj.save()

    def add_vocabulary_base(modeladmin, request, queryset):
        for obj in queryset:
            """"""

    start_trans.short_description = "开始翻译"
    # add_vocabulary_base.short_description = "添加至词汇库"
    actions = [start_trans]


@admin.register(models.VocabularyMatch)
class DataProcess(admin.ModelAdmin):
    # 列表页显示那些字段（列）
    list_display = ["vocabulary", "match_result"]
    search_fields = ["vocabulary"]

    def start_match(modeladmin, request, queryset):
        for obj in queryset:
            vocabulary = obj.vocabulary
            vocabulary_type = obj.vocabulary_type
            is_similarity = obj.is_similarity
            match_result_num = obj.match_result_num

            vm_model_all = models.VocabularyManagement.objects.all()
            vm_all = []
            # vocabulary_type根据类型判断是匹配英文还是中文
            for vm_model in vm_model_all:
                if vocabulary_type == 0:
                    vm_all.append(vm_model.vocabulary_en)
                else:
                    vm_all.append(vm_model.vocabulary_cn)

            # 若使用相似度匹配
            if is_similarity == 0:
                match_threshold = obj.match_threshold
                # 相似度匹配
                match_vm_top = difflib.get_close_matches(
                    vocabulary, vm_all, n=match_result_num, cutoff=match_threshold
                )
                obj.match_result = ",".join(match_vm_top)
            else:
                match_mode = obj.match_mode
                match_vm_top = []
                current_num = 0
                for vm in vm_all:
                    if match_mode == 0:
                        if vocabulary in vm:
                            match_vm_top.append(vm)
                    elif match_mode == 1:
                        if vocabulary == vm:
                            match_vm_top.append(vm)
                    elif match_mode == 2:
                        if str(vm).startswith(vocabulary):
                            match_vm_top.append(vm)
                    elif match_mode == 3:
                        if str(vm).endswith(vocabulary):
                            match_vm_top.append(vm)
                    current_num += 1
                    if current_num == match_result_num:
                        break
                obj.match_result = ",".join(match_vm_top)
            obj.save()

    start_match.short_description = "开始匹配"
    actions = [start_match]


@admin.register(models.VocabularyFileMatch)
class DataProcess(admin.ModelAdmin):
    # 列表页显示那些字段（列）
    list_display = ['id', "vocabulary_file", "result_file_path"]

    def start_match(modeladmin, request, queryset):
        for obj in queryset:
            demand_field_list = []
            field_list = []
            all_field_match_list = []
            remain_field_match_list = []
            final_result_list = []
            no = 0
            book = openpyxl.load_workbook(obj.vocabulary_file, data_only=True)
            sheet = book.worksheets[0]  # 取sheet 页，默认是0，可以根据自己需求进行调整
            for row in sheet.rows:
                row_val = [col.value for col in row]
                print(row_val)
                field_name = row_val[0]
                no += 1
                record_list = [no, field_name]
                print(record_list)
                demand_field_list.append(record_list)
            vm_model_all = models.VocabularyManagement.objects.all()
            for vm_model in vm_model_all:
                dict_record = {vm_model.vocabulary_cn: vm_model.vocabulary_en}
                field_list.append(dict_record)

            for demand_field in demand_field_list:
                for field_remark in field_list:
                    field_name = field_remark.get(demand_field[1])
                    if field_name:
                        match_list = [demand_field[0], demand_field[1]]
                        all_field_match_list.append(match_list)
                        record_list = [demand_field[0], field_name, demand_field[1], '字段匹配']
                        final_result_list.append(record_list)

            no_match_list = [x for x in demand_field_list if x not in all_field_match_list]
            words = ['会计月', '连锁系统']
            jieba.load_userdict(words)
            for df in no_match_list:
                no = df[0]
                df = df[1]
                seg_list = list(jieba.cut(df))
                str = ''
                flag = 0
                str1 = ''
                participle_rule = ''
                for property_field in seg_list:
                    for property_standard in field_list:
                        field_name = property_standard.get(property_field)
                        if field_name:
                            flag += 1
                            str += field_name + '_'
                            str1 += field_name + ':' + property_field + ';'
                    participle_rule += '/' + property_field
                str = str.rstrip('_')
                if flag == len(seg_list):
                    record_list = [no, str, df, '属性匹配', '匹配', participle_rule, str1]
                    all_field_match_list.append(df)
                    final_result_list.append(record_list)
                else:
                    record_list = [no, str, df, '属性匹配', '不匹配', participle_rule, str1]
                    all_field_match_list.append(df)
                    final_result_list.append(record_list)
            final_result_list.sort(key=lambda x: x[0])
            print(final_result_list)

            outwb = Workbook()
            outws = outwb.worksheets[0]
            header = ["序号", "字段名", "字段备注", "匹配类型", "匹配度", "分词规则", "匹配记录"]
            outws.append(header)

            # 设置表头背景为黄色
            fill_yellow = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )
            for cell in outws[1]:  # 第一个单元格是表头
                cell.fill = fill_yellow

            # 设置表头字体加粗
            font_bold = Font(bold=True)
            for cell in outws[1]:  # 第一个单元格是表头
                cell.font = font_bold
            for new_dict in final_result_list:
                outws.append(new_dict)
            outwb.save("DIR_PATH/vocabulary_management/vocabulary_file_dir/2.xlsx")
            obj.result_file_path = "DIR_PATH/vocabulary_management/vocabulary_file_dir/2.xlsx"
            obj.save()

    def get_result_data(modeladmin, request, queryset):
        for obj in queryset:
            result_file_path = obj.result_file_path
            file = open(result_file_path, "rb")
            response = HttpResponse(file)
            response["Content-Type"] = "application/octet-stream"
            response["Content-Disposition"] = 'attachment;filename="{}"'.format(
                result_file_path.encode("utf-8").decode("ISO-8859-1")
            )
            return response

    start_match.short_description = "开始匹配"
    get_result_data.short_description = "获取文件检测结果"

    actions = [start_match, get_result_data]
